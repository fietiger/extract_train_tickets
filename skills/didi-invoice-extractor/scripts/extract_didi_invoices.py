import pdfplumber
import pandas as pd
import os
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def extract_invoice_info(pdf_path):
    info = {
        "文件名": os.path.basename(pdf_path),
        "开票日期": "",
        "金额": 0.0,
        "购买方名称": "",
        "购买方识别号": "",
        "销售方名称": "",
        "销售方识别号": ""
    }
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[0]
            text = page.extract_text()
            
            # Date
            date_match = re.search(r"开票日期[:：]\s*(\d{4}年\d{1,2}月\d{1,2}日)", text)
            if date_match:
                info["开票日期"] = date_match.group(1)
                
            # Amount
            lines = text.split('\n')
            for line in lines:
                if "价税合计" in line:
                    amt_match = re.search(r"¥\s*([\d\.]+)", line)
                    if amt_match:
                        info["金额"] = float(amt_match.group(1))
                        break
            
            if info["金额"] == 0.0:
                amts = re.findall(r"¥\s*([\d\.]+)", text)
                if amts:
                    info["金额"] = float(amts[-1])

            # Names and Tax IDs
            name_matches = re.findall(r"名称[:：]\s*([^\n\s]+)", text)
            if len(name_matches) >= 1:
                info["购买方名称"] = name_matches[0].strip()
            if len(name_matches) >= 2:
                info["销售方名称"] = name_matches[1].strip()
                
            tax_matches = re.findall(r"纳税人识别号[:：]\s*([A-Z0-9]+)", text)
            if len(tax_matches) >= 1:
                info["购买方识别号"] = tax_matches[0].strip()
            if len(tax_matches) >= 2:
                info["销售方识别号"] = tax_matches[1].strip()
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")

    return info

def process_directory(input_dir, output_file):
    if not os.path.exists(input_dir):
        print(f"Error: Directory '{input_dir}' does not exist.")
        return

    results = []
    pdf_files = [f for f in os.listdir(input_dir) if f.endswith('.pdf') and '发票' in f]
    
    if not pdf_files:
        print(f"No Didi invoice PDF files found in '{input_dir}'.")
        return

    for f in pdf_files:
        path = os.path.join(input_dir, f)
        print(f"Processing: {f}")
        results.append(extract_invoice_info(path))

    if not results:
        print("No information extracted.")
        return

    # Create Excel with openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "发票汇总"

    # Headers
    headers = ["文件名", "开票日期", "金额", "购买方名称", "购买方识别号", "销售方名称", "销售方识别号"]
    ws.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Add data
    for r in results:
        ws.append([
            r["文件名"],
            r["开票日期"],
            f'={r["金额"]}',
            r["购买方名称"],
            f'="{r["购买方识别号"]}"',
            r["销售方名称"],
            f'="{r["销售方识别号"]}"'
        ])


    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file)
    print(f"Success! Saved {len(results)} invoices to: {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python extract_didi_invoices.py <input_dir> <output_xlsx>")
        sys.exit(1)
    
    process_directory(sys.argv[1], sys.argv[2])
