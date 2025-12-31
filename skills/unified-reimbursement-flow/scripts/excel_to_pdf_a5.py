import os
import win32com.client
import sys
import argparse
import pdfplumber
import re
from openpyxl import load_workbook

def convert_to_pdf_win32(xlsx_path, pdf_path):
    abs_xlsx = os.path.abspath(xlsx_path)
    abs_pdf = os.path.abspath(pdf_path)
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(abs_xlsx)
        for ws in wb.Worksheets:
            ws.PageSetup.PaperSize = 11 # A5
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = False
        wb.ExportAsFixedFormat(0, abs_pdf)
        wb.Close(False)
    finally:
        excel.Quit()

def update_count(xlsx_path, pdf_path):
    if not os.path.exists(pdf_path): return
    with pdfplumber.open(pdf_path) as pdf:
        n = len(pdf.pages)
    if n > 1:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        val = str(ws['J3'].value)
        match = re.search(r'(\d+)', val)
        if match:
            new_count = int(match.group(1)) + (n - 1)
            ws['J3'] = f"单据及附件共{new_count}页"
            wb.save(xlsx_path)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx")
    parser.add_argument("pdf")
    parser.add_argument("--update-count", action="store_true")
    args = parser.parse_args()
    
    # Special logic for reimbursement form: 
    # The requirement was to check page count of '费用清单.pdf' and update '费用报销单.xlsx'
    # Here we adapt: if --update-count is passed, we assume we are converting '费用报销单' 
    # and we should check '费用清单.pdf' first.
    if args.update_count and os.path.exists('费用清单.pdf'):
        with pdfplumber.open('费用清单.pdf') as pdf:
            extra = len(pdf.pages) - 1
            if extra > 0:
                wb = load_workbook(args.xlsx)
                ws = wb.active
                match = re.search(r'(\d+)', str(ws['J3'].value))
                if match:
                    ws['J3'] = f"单据及附件共{int(match.group(1)) + extra}页"
                    wb.save(args.xlsx)

    convert_to_pdf_win32(args.xlsx, args.pdf)
