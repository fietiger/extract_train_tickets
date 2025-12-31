import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from copy import copy
import datetime
import os

def generate_expense_list():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'assets', 'expense_template.xlsx')
    train_path = '火车票汇总信息表.xlsx'
    didi_path = '滴滴行程明细汇总表.xlsx'
    output_path = '费用清单.xlsx'


    consolidated_data = []
    if os.path.exists(train_path):
        df_train = pd.read_excel(train_path)
        for _, row in df_train.iterrows():
            price = str(row['price']).replace('=', '').replace('¥', '').replace(',', '').strip()
            consolidated_data.append({'日期': str(row['date']), '事由': f"出差交通({row['departure_station']}-{row['arrival_station']})", '项目名称': '公共项目', '类别': '长途交通费', '金额': float(price), '备注': ''})

    if os.path.exists(didi_path):
        df_didi = pd.read_excel(didi_path)
        for _, row in df_didi.iterrows():
            time_str = str(row['上车时间'])
            try:
                dt = datetime.datetime.strptime(f"2025-{time_str}", '%Y-%m-%d %H:%M')
                date_str = dt.strftime('%Y-%m-%d')
            except: date_str = time_str
            consolidated_data.append({'日期': date_str, '事由': '市内交通', '项目名称': '公共项目', '类别': '市内交通费', '金额': row['金额[元]'], '备注': ''})

    df_result = pd.DataFrame(consolidated_data).sort_values(by='日期')
    wb = load_workbook(template_path)
    ws = wb.active
    while ws.max_row >= 4: ws.delete_rows(4)
    for _, row_data in df_result.iterrows():
        ws.append([row_data['日期'], row_data['事由'], row_data['项目名称'], row_data['类别'], row_data['金额'], row_data['备注']])
    
    current_row = ws.max_row + 1
    ws.cell(row=current_row, column=1, value='合计')
    ws.cell(row=current_row, column=5, value=f"=SUM(E4:E{current_row-1})")
    wb.save(output_path)

if __name__ == "__main__":
    generate_expense_list()
