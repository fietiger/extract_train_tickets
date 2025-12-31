import pdfplumber
import pandas as pd
import os
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def extract_invoice_info(pdf_path):
    info = {"文件名": os.path.basename(pdf_path), "开票日期": "", "金额": 0.0, "购买方名称": "", "购买方识别号": "", "销售方名称": "", "销售方识别号": ""}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text()
            date_match = re.search(r"开票日期[:：]\s*(\d{4}年\d{1,2}月\d{1,2}日)", text)
            if date_match: info["开票日期"] = date_match.group(1)
            amts = re.findall(r"¥\s*([\d\.]+)", text)
            if amts: info["金额"] = float(amts[-1])
            name_matches = re.findall(r"名称[:：]\s*([^\n\s]+)", text)
            if len(name_matches) >= 1: info["购买方名称"] = name_matches[0].strip()
            if len(name_matches) >= 2: info["销售方名称"] = name_matches[1].strip()
            tax_matches = re.findall(r"纳税人识别号[:：]\s*([A-Z0-9]+)", text)
            if len(tax_matches) >= 1: info["购买方识别号"] = tax_matches[0].strip()
            if len(tax_matches) >= 2: info["销售方识别号"] = tax_matches[1].strip()
    except: pass
    return info

def process_directory(input_dir, output_file):
    pdf_files = [f for f in os.listdir(input_dir) if f.endswith('.pdf') and '发票' in f]
    results = [extract_invoice_info(os.path.join(input_dir, f)) for f in pdf_files]
    if not results: return
    wb = Workbook()
    ws = wb.active
    headers = ["文件名", "开票日期", "金额", "购买方名称", "购买方识别号", "销售方名称", "销售方识别号"]
    ws.append(headers)
    for r in results:
        ws.append([r["文件名"], r["开票日期"], f'={r["金额"]}', r["购买方名称"], f'="{r["购买方识别号"]}"', r["销售方名称"], f'="{r["销售方识别号"]}"'])
    wb.save(output_file)

if __name__ == "__main__":
    process_directory(sys.argv[1], sys.argv[2])
