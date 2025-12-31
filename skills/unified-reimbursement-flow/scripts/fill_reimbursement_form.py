import pandas as pd
from openpyxl import load_workbook
import datetime
import os

def count_pdfs(directory):
    count = 0
    if not os.path.exists(directory): return 0
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.pdf'): count += 1
    return count

def fill_reimbursement():
    train_file, didi_file = '火车票.xlsx', '滴滴发票.xlsx'
    train_sum = 0
    if os.path.exists(train_file):
        df = pd.read_excel(train_file)
        train_sum = df['price'].apply(lambda x: float(str(x).replace('=', '')) if pd.notna(x) else 0).sum()
    
    didi_sum = 0
    if os.path.exists(didi_file):
        df = pd.read_excel(didi_file)
        didi_sum = df['金额'].apply(lambda x: float(str(x).replace('=', '')) if pd.notna(x) else 0).sum()

    total_pages = count_pdfs('滴滴出行电子发票及行程报销单') + count_pdfs('火车票') + 2
    
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_file = os.path.join(base_dir, 'assets', 'reimbursement_template.xlsx')
    wb = load_workbook(template_file)
    ws = wb.active
    ws['E5'] = train_sum + didi_sum
    ws['E6'], ws['E7'] = 0, 0
    now = datetime.datetime.now()
    ws['D3'] = f"{now.year} 年 {now.month}月{now.day} 日 填"
    ws['J3'] = f"单据及附件共{total_pages}页"
    wb.save('费用报销单.xlsx')

if __name__ == "__main__":
    fill_reimbursement()
