import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from copy import copy
import datetime
import os

def generate_expense_list():
    # 1. 定义文件路径 (使用相对路径或从 skill 资源目录读取)
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'assets', 'expense_template.xlsx')
    
    # 输入文件默认在当前工作目录
    train_path = os.path.abspath('火车票提取结果_最新.xlsx')
    didi_path = os.path.abspath('滴滴行程汇总_技能执行版.xlsx')
    output_path = os.path.abspath('费用清单_最新版.xlsx')


    # 2. 读取火车票数据 (使用 openpyxl 避免公式读取问题)
    wb_train = load_workbook(train_path)
    ws_train = wb_train.active
    train_data = []
    header = [cell.value for cell in ws_train[1]]
    for row in ws_train.iter_rows(min_row=2):
        row_dict = {header[i]: cell.value for i, cell in enumerate(row)}
        train_data.append(row_dict)

    def clean_price(val):
        if val is None: return 0.0
        # 清理 '=' '¥' ',' 等符号
        s = str(val).replace('=', '').replace('¥', '').replace(',', '').strip()
        try:
            return float(s)
        except:
            return 0.0

    # 3. 整理汇总数据
    consolidated_data = []

    # 处理火车票
    for row in train_data:
        price = clean_price(row['price'])
        dep = str(row['departure_station'])
        arr = str(row['arrival_station'])
        dt_val = row['date']
        date_str = dt_val.strftime('%Y-%m-%d') if isinstance(dt_val, (datetime.datetime, datetime.date)) else str(dt_val)
        
        consolidated_data.append({
            '日期': date_str,
            '事由': f"出差交通({dep}-{arr})",
            '项目名称': '公共项目',
            '类别': '长途交通费',
            '金额': price,
            '备注': ''
        })

    # 处理滴滴行程
    df_didi = pd.read_excel(didi_path)
    for _, row in df_didi.iterrows():
        time_str = str(row['上车时间'])
        try:
            # 补全年份为 2025
            dt = datetime.datetime.strptime(f"2025-{time_str}", '%Y-%m-%d %H:%M')
            date_str = dt.strftime('%Y-%m-%d')
        except:
            date_str = time_str
        
        consolidated_data.append({
            '日期': date_str,
            '事由': '市内交通',
            '项目名称': '公共项目',
            '类别': '市内交通费',
            '金额': row['金额[元]'],
            '备注': ''
        })

    # 按日期排序
    df_result = pd.DataFrame(consolidated_data).sort_values(by='日期')

    # 4. 写入模板并设置格式
    wb = load_workbook(template_path)
    ws = wb.active

    # 提取第 4 行或第 3 行作为样式模板
    sample_row = ws[4] if ws.max_row >= 4 else ws[3]
    styles = []
    for cell in sample_row:
        styles.append({
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'alignment': copy(cell.alignment)
        })

    # 清除原有数据 (从第 4 行开始)
    while ws.max_row >= 4:
        ws.delete_rows(4)

    # 货币格式 (¥ 符号 + 2位小数)
    rmb_format = '¥#,##0.00'

    # 填充新数据
    current_row = 4
    for _, row_data in df_result.iterrows():
        ws.append([
            row_data['日期'],
            row_data['事由'],
            row_data['项目名称'],
            row_data['类别'],
            row_data['金额'],
            row_data['备注']
        ])
        
        # 应用样式
        for i, cell in enumerate(ws[current_row]):
            if i < len(styles):
                cell.font = copy(styles[i]['font'])
                cell.border = copy(styles[i]['border'])
                cell.fill = copy(styles[i]['fill'])
                cell.alignment = copy(styles[i]['alignment'])
            if i == 4: # 金额列 (E)
                cell.number_format = rmb_format
        current_row += 1

    # 添加合计行
    ws.cell(row=current_row, column=1, value='合计')
    for i, cell in enumerate(ws[current_row]):
        if i < len(styles):
            cell.font = copy(styles[i]['font'])
            cell.border = copy(styles[i]['border'])
            cell.fill = copy(styles[i]['fill'])
            cell.alignment = copy(styles[i]['alignment'])
        if i == 4: # 金额列合计
            cell.value = f"=SUM(E4:E{current_row-1})"
            cell.number_format = rmb_format

    # 5. 特殊处理 A2:F2 合并单元格边框 (移除整个范围的左右外边框)
    # 对于合并单元格，需要处理范围边界上的所有单元格
    for col in range(1, 7): # A to F
        cell = ws.cell(row=2, column=col)
        current_border = cell.border
        left_side = Side(style=None) if col == 1 else current_border.left
        right_side = Side(style=None) if col == 6 else current_border.right
        
        cell.border = Border(
            left=left_side,
            right=right_side,
            top=current_border.top,
            bottom=current_border.bottom
        )


    # 6. 保存结果
    wb.save(output_path)
    print(f"成功生成最终费用清单: {output_path}")

if __name__ == "__main__":
    generate_expense_list()
