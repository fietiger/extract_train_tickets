import pandas as pd
from openpyxl import load_workbook
import datetime
import os

def count_pdfs(directory):
    count = 0
    if not os.path.exists(directory):
        return 0
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.pdf'):
                count += 1
    return count

def fill_reimbursement():
    # 1. 汇总火车票金额
    train_file = '火车票汇总信息表.xlsx'
    train_sum = 0
    if os.path.exists(train_file):
        wb_train = load_workbook(train_file, data_only=False)
        ws_train = wb_train.active
        # 假设价格在第12列 (L列)，从第2行开始
        for row in range(2, ws_train.max_row + 1):
            val = ws_train.cell(row=row, column=12).value
            if val:
                if isinstance(val, str) and val.startswith('='):
                    try:
                        train_sum += float(val[1:])
                    except: pass
                else:
                    try:
                        train_sum += float(val)
                    except: pass
    
    # 2. 汇总滴滴发票金额
    didi_file = '滴滴电子发票汇总.xlsx'
    didi_sum = 0
    if os.path.exists(didi_file):
        wb_didi = load_workbook(didi_file, data_only=False)
        ws_didi = wb_didi.active
        
        # 查找“金额”列的索引
        amount_col = -1
        for col in range(1, ws_didi.max_column + 1):
            if ws_didi.cell(row=1, column=col).value == '金额':
                amount_col = col
                break
        
        if amount_col != -1:
            for row in range(2, ws_didi.max_row + 1):
                val = ws_didi.cell(row=row, column=amount_col).value
                if val:
                    if isinstance(val, str) and val.startswith('='):
                        try:
                            didi_sum += float(val[1:])
                        except: pass
                    else:
                        try:
                            didi_sum += float(val)
                        except: pass


    total_transport = train_sum + didi_sum
    
    # 3. 统计 PDF 数量
    didi_pdf_count = count_pdfs('滴滴出行电子发票及行程报销单')
    train_pdf_count = count_pdfs('火车票')
    total_pages = didi_pdf_count + train_pdf_count + 2
    
    # 4. 填充模板
    # 获取脚本所在目录，以便定位 assets 文件夹
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_file = os.path.join(script_dir, '..', 'assets', '费用报销单模板.xlsx')
    
    if not os.path.exists(template_file):
        # 降级处理：尝试读取当前目录下的模板
        template_file = '费用报销单模板.xlsx'
        if not os.path.exists(template_file):
            print(f"Error: Template {template_file} not found.")
            return

    wb = load_workbook(template_file)

    ws = wb.active
    
    # 填写交通费
    ws['E5'] = total_transport
    ws['E6'] = 0
    ws['E7'] = 0
    
    # 填写日期
    now = datetime.datetime.now()
    ws['D3'] = f"{now.year} 年 {now.month}月{now.day} 日 填"
    
    # 填写单据页数
    ws['J3'] = f"单据及附件共{total_pages}页"
    
    # 保存结果
    output_file = '费用报销单.xlsx'
    wb.save(output_file)
    print(f"Success: {output_file} generated with total {total_transport} and {total_pages} pages.")

if __name__ == "__main__":
    fill_reimbursement()
